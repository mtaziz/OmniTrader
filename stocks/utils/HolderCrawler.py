import scrapy
from stocks.models import *
from scrapy.http import Request
import re
import xlrd
from openpyxl import load_workbook


class FlowValueItem(scrapy.Item):
    ticker = scrapy.Field()
    row = scrapy.Field()
    flow_value = scrapy.Field()
    total_value = scrapy.Field()
    pct = scrapy.Field()

class FlowValueItemPipeline(object):

    FLOW_VALUE_COL = 6
    TOTAL_VALUE_COL = 5

    def open_spider(self, spider):
        self.wb = load_workbook(filename=spider.filepath)

    def close_spider(self, spider):
        self.wb.save(spider.filepath)


    def process_item(self, item, spider):
        sheet = self.wb.worksheets[0]
        flow_value = round(item['flow_value']*item['pct']/100, 1)
        total_value = round(item['total_value'], 1)
        sheet.cell(row=item['row'], column=self.FLOW_VALUE_COL).value = flow_value
        sheet.cell(row=item['row'], column=self.TOTAL_VALUE_COL).value = total_value




class HolderCrawler(scrapy.Spider):
    name = "HolderCrawler"
    allowed_domains = ["10jqka.com.cn"]
    start_urls = [
        "http://stockpage.10jqka.com.cn/600340/",
    ]



    def parseHolder(self,response):
        item = response.request.meta['item']

        #Trim the url to get the stock code
        #ticker = response.url[31:37]

        content = response.xpath('//div[@id="fher_1"]//tbody/tr')
        accumulated_pct = 0.0
        for tr in content:
            holder = tr.xpath('th/a/text()').extract()[0]
            raw_pct = tr.xpath('td[3]/text()').extract()[0]
            try:
                # Deal with '-' aka unknown value.
                if raw_pct.startswith('-'):
                    break
                # Deal with occassional pct sign at the end
                if raw_pct.endswith('%'):
                    raw_pct = raw_pct[:-1]
                pct = float(raw_pct)
                if pct >= 5:
                    accumulated_pct += pct
                else:
                    # Holders are listed by share pct order in THS. if less than 5% then no need to go on.
                    break
            except ValueError:
                break

        item['pct'] = 100-accumulated_pct

        yield Request("http://d.10jqka.com.cn/v2/realhead/hs_{}/last.js".format(item['ticker']), meta={'item': item}, callback=self.parseBasic)


    def parseBasic(self,response):
        # A sample response is as follow from request http://d.10jqka.com.cn/v2/realhead/hs_600340/last.js
        # text = r'quotebridge_v2_realhead_hs_600340_last({"items":{"402":"2954946709.00","407":"2954946709.00","527198":"52579056.000","10":"35.18","24":"35.16","25":"17500.00","30":"35.17","31":"1200.00","8":"35.95","9":"34.51","13":"94672636.00","19":"3347188400.00","7":"35.00","15":"45764725.00","14":"42093580.00","69":"38.42","70":"31.44","223":"398724910.00","224":"556370570.00","225":"495770220.00","226":"407028100.00","237":"277437830.00","238":"229685520.00","259":"545106020.00","260":"437078830.00","38":"-1","37":"-1","39":"-1","23":"","22":"","90":"","92":"","254":"","278":"","49":"1000.00","271":"-1","51":"","276":"","277":"","12":"5","17":"962900.00","95":"","96":"","273":"-1","274":"","264648":"0.250","199112":"0.716","1968584":"3.204","2034120":"16.014","1378761":"35.355","526792":"4.123","395720":"29723.000","461256":"18.947","3475914":"103955025000.000","3541450":"103955025000.000","1771976":"0.430","134152":"16.014","592920":"4.099","6":"34.93","66":"","stop":0,"time":"2017-04-24 20:45:41 \u5317\u4eac\u65f6\u95f4","name":"\u534e\u590f\u5e78\u798f","marketid":"17"}})'

        item = response.request.meta['item']
        # Flow value is after 3475914 tag. Hard coding here.
        m = re.search('3475914\"\:\"(\d+\.\d+)\".+3541450\"\:\"(\d+\.\d+)\"', str(response.body))
        if m:
            item['flow_value'] = float(m.group(1))/100000000
            item['total_value'] = float(m.group(2))/100000000

        yield item



    def parse(self,response):
        self.workbook = xlrd.open_workbook(self.filepath)
        self.worksheet = self.workbook.sheet_by_index(0)

        for i in range(0, self.worksheet.nrows):
            ticker = self.worksheet.cell(i, 0).value
            if re.fullmatch('\d{6}\Z', ticker):
                item = FlowValueItem()
                item['ticker'] = ticker
                item['row'] = i+1
                yield Request("http://stockpage.10jqka.com.cn/{}/holder".format(ticker), meta={'item': item}, callback=self.parseHolder)
